import openpyxl


def copy_first_column(input_file, output_file):
    # Wczytanie pliku wejściowego
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active

    # Utworzenie nowego pliku wyjściowego
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active

    # Pobranie wartości z pierwszej kolumny
    column_values = [cell.value for cell in ws_input['A']]

    # Zapisanie wartości do nowego pliku w pierwszej kolumnie
    for index, value in enumerate(column_values, start=1):
        ws_output.cell(row=index, column=1, value=value)

    # Zapisanie zmian do pliku wyjściowego
    wb_output.save(output_file)

    print("Pierwsza kolumna została zapisana do pliku", output_file)


# Nazwy plików wejściowego i wyjściowego
input_file_name = 'output_ch.xlsx'
output_file_name = 'output_ch2.xlsx'

# Wywołanie funkcji do kopiowania pierwszej kolumny
copy_first_column(input_file_name, output_file_name)
