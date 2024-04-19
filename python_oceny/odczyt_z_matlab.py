import openpyxl

def read_row_values(sheet, row_number):
    """Funkcja do odczytywania wartości wiersza z arkusza."""
    row_values = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row_number, column=col).value
        row_values.append(cell_value)
    return row_values

def write_column_values(sheet, col_number, values):
    """Funkcja do zapisywania wartości do kolumny w arkuszu."""
    for row, value in enumerate(values, start=1):
        sheet.cell(row=row, column=col_number, value=value)

# Wczytanie pliku wejściowego
input_file_name = 'numery_ch.xlsx'
output_file_name = 'output_ch.xlsx'

wb_input = openpyxl.load_workbook(input_file_name)
ws_input = wb_input.active

# Utworzenie nowego pliku wyjściowego
wb_output = openpyxl.Workbook()
ws_output = wb_output.active

# Odczytanie wartości z wiersza i zapisanie ich do kolejnych kolumn
for col in range(1, 1170):  # 1169 kolumn
    row_values = read_row_values(ws_input, row_number=1)
    write_column_values(ws_output, col, row_values)

# Zapisanie zmian do pliku wyjściowego
wb_output.save(output_file_name)

# Komunikat o zakończeniu
print("Dane zostały pomyślnie zapisane do pliku", output_file_name)
